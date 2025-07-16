#!/usr/bin/env python3

import os
import re
from pathlib import Path
import easyocr
import fitz  # PyMuPDF
from openpyxl import Workbook
from concurrent.futures import ThreadPoolExecutor, as_completed

def read_keywords(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        return [line.strip() for line in f if line.strip()]

def ocr_pdf(pdf_path, reader):
    text = ""
    try:
        pdf_document = fitz.open(pdf_path)
        for page_num in range(pdf_document.page_count):
            page = pdf_document[page_num]
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # 2x scaling for better OCR
            img_data = pix.tobytes("png")
            
            results = reader.readtext(img_data)
            for (bbox, extracted_text, confidence) in results:
                text += extracted_text + " "
        pdf_document.close()
    except Exception as e:
        print(f"Error processing {pdf_path}: {e}")
    
    text = re.sub(r'\s+', ' ', text)  # Replace multiple whitespace with single space
    text = text.strip()  # Remove leading/trailing whitespace
    
    return text

def search_keywords_in_text(text, keywords):
    # Convert text to lowercase for case-insensitive search
    text_lower = text.lower()
    results = {}
    
    for keyword in keywords:
        keyword_lower = keyword.lower()
        
        # Check if keyword contains wildcards
        if '*' in keyword_lower:
            # Convert wildcard pattern to regex pattern
            regex_pattern = keyword_lower.replace('*', '.*')
            # Use word boundaries to match whole words when no wildcards at edges
            if not keyword_lower.startswith('*'):
                regex_pattern = r'\b' + regex_pattern
            if not keyword_lower.endswith('*'):
                regex_pattern = regex_pattern + r'\b'
            
            matches = re.findall(regex_pattern, text_lower)
            results[keyword] = len(matches) > 0
        else:
            # Exact keyword search with word boundaries for case-insensitive match
            pattern = r'\b' + re.escape(keyword_lower) + r'\b'
            matches = re.findall(pattern, text_lower)
            results[keyword] = len(matches) > 0
    
    return results

def create_excel_report(results, keywords, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Keyword Search Results"
    
    # Headers
    ws['A1'] = 'PDF File'
    for i, keyword in enumerate(keywords, start=2):
        ws.cell(row=1, column=i, value=keyword)
    
    # Data
    for row, (pdf_file, keyword_results) in enumerate(results.items(), start=2):
        ws.cell(row=row, column=1, value=pdf_file)
        for col, keyword in enumerate(keywords, start=2):
            ws.cell(row=row, column=col, value='YES' if keyword_results[keyword] else 'NO')
    
    wb.save(output_path)

def process_single_pdf(pdf_file, keywords):
    """Process a single PDF file - for parallel execution"""
    # Create a separate OCR reader for each thread to avoid conflicts
    reader = easyocr.Reader(['en'])
    
    print(f"Processing {pdf_file.name}...")
    text = ocr_pdf(pdf_file, reader)
    keyword_results = search_keywords_in_text(text, keywords)
    return pdf_file.name, keyword_results

def main():
    # Setup paths
    keywords_file = 'keywords.txt'
    inputs_dir = Path('inputs')
    outputs_dir = Path('outputs')
    
    # Create outputs directory if it doesn't exist
    outputs_dir.mkdir(exist_ok=True)
    
    # Read keywords
    if not os.path.exists(keywords_file):
        print(f"Error: {keywords_file} not found")
        return
    
    keywords = read_keywords(keywords_file)
    if not keywords:
        print("No keywords found")
        return
    
    # Process PDFs
    pdf_files = list(inputs_dir.glob('*.pdf'))
    if not pdf_files:
        print("No PDF files found in inputs/ directory")
        return
    
    results = {}
    
    # Use parallel processing to speed up PDF processing
    # Limit max workers to avoid resource conflicts
    max_workers = min(4, len(pdf_files))
    
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # Submit all PDF processing tasks
        future_to_pdf = {executor.submit(process_single_pdf, pdf_file, keywords): pdf_file 
                        for pdf_file in pdf_files}
        
        # Collect results as they complete
        for future in as_completed(future_to_pdf):
            try:
                pdf_name, keyword_results = future.result()
                results[pdf_name] = keyword_results
            except Exception as e:
                pdf_file = future_to_pdf[future]
                print(f"Error processing {pdf_file.name}: {e}")
    
    # Create Excel report
    output_file = outputs_dir / 'keyword_search_results.xlsx'
    create_excel_report(results, keywords, output_file)
    
    print(f"Report saved to {output_file}")

if __name__ == "__main__":
    main()